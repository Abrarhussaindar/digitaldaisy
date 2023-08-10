from django.contrib import admin
from .models import *
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.http import HttpResponse
import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import timedelta


class MyUserAdmin(BaseUserAdmin):
    list_display = ('username', 'empid', 'permission','work_time', 'department', 'intime', 'out_time', 'duration_time', 'is_superuser', 'is_admin', 'is_teamLead', 'is_staff', 'is_active')
    search_fields = ('username', 'email')
    filter_horizontal = ()
    list_filter = ('last_login',)
    fieldsets = ()

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('name', 'empid', 'permission', 'department', 'designation', 'phone_number', 'login_counter', 'logout_counter', 'intime', 'out_time', 'duration_time', 'username', 'email', 'password1', 'password2'),
        }),
    )
    ordering = ('name', 'username', 'email',)


def parse_duration_time(duration_time_str):
    if duration_time_str:
        duration_parts = duration_time_str.split(':')
        hours = int(float(duration_parts[0]))
        minutes = int(float(duration_parts[1]))
        seconds = int(float(duration_parts[2]))
        return timedelta(hours=hours, minutes=minutes, seconds=seconds)
    return None


from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment
from io import BytesIO
from datetime import datetime

def download_daily_employee_details(modeladmin, request, queryset):
    # Define the header row
    header_row = [
        "Employee Name", "Username",   "Department", "Designation","Work Time", "Remaining Time","Login Time", "Logout Time", "Duration",
        "Login List", "Logout List", "Duration List"
    ]

    # Create a new workbook
    workbook = Workbook()

    # Create a new worksheet
    worksheet = workbook.active
    worksheet.title = "Daily Employees Attendance"

    # Set column widths and alignment
    for col_num, column_title in enumerate(header_row, 1):
        worksheet.column_dimensions[chr(64 + col_num)].width = 15
        worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')

    # Write the header row
    worksheet.append(header_row)

    # Filter employees based on departments (excluding "Management/Admin" department)
    departments = set(queryset.exclude(department="Management/Admin").values_list('department', flat=True))
    for department in departments:
        filtered_employees = queryset.filter(department=department)

        # Write the data rows for each department
        for employee in filtered_employees:
            login_list = employee.login_list if employee.login_list else []
            logout_list = employee.logout_list if employee.logout_list else []
            duration_list = employee.duration_list if employee.duration_list else []
        
            total_work_time = employee.work_time
            remaining_time = employee.remaining_time

            # Get the current row number
            row_num = worksheet.max_row + 1

            # Write the data to the worksheet
            worksheet.cell(row=row_num, column=1).value = employee.name
            worksheet.cell(row=row_num, column=2).value = employee.username
            worksheet.cell(row=row_num, column=3).value = employee.department
            worksheet.cell(row=row_num, column=4).value = employee.designation
            worksheet.cell(row=row_num, column=5).value = total_work_time
            worksheet.cell(row=row_num, column=6).value = remaining_time
            worksheet.cell(row=row_num, column=7).value = employee.intime
            worksheet.cell(row=row_num, column=8).value = employee.out_time
            worksheet.cell(row=row_num, column=9).value = employee.duration_time
            worksheet.cell(row=row_num, column=10).value = "\n".join(login_list)
            worksheet.cell(row=row_num, column=11).value = "\n".join(logout_list)
            worksheet.cell(row=row_num, column=12).value = "\n".join(duration_list)

    excel_file = BytesIO()

    # Save the workbook to the BytesIO object
    workbook.save(excel_file)
    date = datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%d:%m:%Y')
    file_name = f"{date}_employees_attendance.xlsx"
    print("file_name: ", file_name)
    
    # Set the response headers for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    # Write the BytesIO data to the response
    response.write(excel_file.getvalue())

    return response



download_daily_employee_details.short_description = "Download Daily employees Attendance as Excel"

def download_employee_details(modeladmin, request, queryset):
    # Define the header row
    header_row = [
        "Employee Name", "Username",   "Department", "Designation","Work Time", "Remaining Time","Login Time", "Logout Time", "Duration",
        "Login List", "Logout List", "Duration List",
        "Weekly Login List", "Weekly Logout List", "Weekly Duration List",
        "Monthly Login List", "Monthly Logout List", "Monthly Duration List",
        "Yearly Login List", "Yearly Logout List", "Yearly Duration List", "Yearly Login Dates"
    ]

    # Create a new workbook
    workbook = Workbook()

    # Create a new worksheet
    worksheet = workbook.active
    worksheet.title = "Employee Attendance"

    # Set column widths and alignment
    for col_num, column_title in enumerate(header_row, 1):
        worksheet.column_dimensions[chr(64 + col_num)].width = 15
        worksheet.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')

    # Write the header row
    worksheet.append(header_row)

    # Filter employees based on departments (excluding "Management/Admin" department)
    departments = set(queryset.exclude(department="Management/Admin").values_list('department', flat=True))
    for department in departments:
        filtered_employees = queryset.filter(department=department)

        # Write the data rows for each department
        for employee in filtered_employees:
            login_list = employee.login_list if employee.login_list else []
            logout_list = employee.logout_list if employee.logout_list else []
            duration_list = employee.duration_list if employee.duration_list else []
            weekly_login_list = employee.weekly_login_list if employee.weekly_login_list else []
            weekly_logout_list = employee.weekly_logout_list if employee.weekly_logout_list else []
            weekly_duration_list = employee.weekly_duration_list if employee.weekly_duration_list else []
            monthly_login_list = employee.monthly_login_list if employee.monthly_login_list else []
            monthly_logout_list = employee.monthly_logout_list if employee.monthly_logout_list else []
            monthly_duration_list = employee.monthly_duration_list if employee.monthly_duration_list else []
            yearly_login_list = employee.yearly_login_list if employee.yearly_login_list else []
            yearly_logout_list = employee.yearly_logout_list if employee.yearly_logout_list else []
            yearly_duration_list = employee.yearly_duration_list if employee.yearly_duration_list else []
            yearly_login_dates = employee.yearly_login_dates if employee.yearly_login_dates else []

            # Convert datetime objects to timezone-aware objects with tzinfo set to None
            login_time = employee.login_time.astimezone().replace(tzinfo=None) if employee.login_time else None
            logout_time = employee.logout_time.astimezone().replace(tzinfo=None) if employee.logout_time else None
            duration_time = parse_duration_time(employee.duration_time)

            # Get the total work time and remaining time
            total_work_time = employee.work_time
            remaining_time = employee.remaining_time

            # Get the current row number
            row_num = worksheet.max_row + 1

            # Write the data to the worksheet
            worksheet.cell(row=row_num, column=1).value = employee.name
            worksheet.cell(row=row_num, column=2).value = employee.username
            worksheet.cell(row=row_num, column=3).value = employee.department
            worksheet.cell(row=row_num, column=4).value = employee.designation
            worksheet.cell(row=row_num, column=5).value = total_work_time
            worksheet.cell(row=row_num, column=6).value = remaining_time
            worksheet.cell(row=row_num, column=7).value = employee.intime
            worksheet.cell(row=row_num, column=8).value = employee.out_time
            worksheet.cell(row=row_num, column=9).value = employee.duration_time
            worksheet.cell(row=row_num, column=10).value = "\n".join(login_list)
            worksheet.cell(row=row_num, column=11).value = "\n".join(logout_list)
            worksheet.cell(row=row_num, column=12).value = "\n".join(duration_list)
            worksheet.cell(row=row_num, column=13).value = "\n".join(weekly_login_list)
            worksheet.cell(row=row_num, column=14).value = "\n".join(weekly_logout_list)
            worksheet.cell(row=row_num, column=15).value = "\n".join(weekly_duration_list)
            worksheet.cell(row=row_num, column=16).value = "\n".join(monthly_login_list)
            worksheet.cell(row=row_num, column=17).value = "\n".join(monthly_logout_list)
            worksheet.cell(row=row_num, column=18).value = "\n".join(monthly_duration_list)
            worksheet.cell(row=row_num, column=19).value = "\n".join(yearly_login_list)
            worksheet.cell(row=row_num, column=20).value = "\n".join(yearly_logout_list)
            worksheet.cell(row=row_num, column=21).value = "\n".join(yearly_duration_list)
            worksheet.cell(row=row_num, column=22).value = "\n".join(yearly_login_dates)

    excel_file = BytesIO()

    # Save the workbook to the BytesIO object
    workbook.save(excel_file)

    # Set the response headers for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_attendance_details.xlsx"'

    # Write the BytesIO data to the response
    response.write(excel_file.getvalue())

    return response

download_employee_details.short_description = "Download selected employee details as Excel"
# Register the Employee model with the admin site
admin.site.register(Employee, MyUserAdmin)
admin.site.add_action(download_employee_details)
admin.site.register(IndividualEmployeeAttendance)

